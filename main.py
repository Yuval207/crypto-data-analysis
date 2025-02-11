import requests
import pandas as pd
import time
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Constants
API_URL = "https://api.coingecko.com/api/v3/coins/markets"
PARAMS = {
    "vs_currency": "usd",
    "order": "market_cap_desc",
    "per_page": 50,
    "page": 1,
    "sparkline": False
}
EXCEL_FILE = "crypto_live_data.xlsx"
UPDATE_INTERVAL = 300  # 5 minutes

def fetch_crypto_data():
    response = requests.get(API_URL, params=PARAMS)
    if response.status_code == 200:
        return response.json()
    else:
        print("Failed to fetch data", response.status_code)
        return []

def analyze_data(data):
    df = pd.DataFrame(data)
    top_5 = df.nlargest(5, 'market_cap')[['name', 'market_cap']]
    avg_price = df['current_price'].mean()
    highest_change = df.nlargest(1, 'price_change_percentage_24h')[['name', 'price_change_percentage_24h']]
    lowest_change = df.nsmallest(1, 'price_change_percentage_24h')[['name', 'price_change_percentage_24h']]
    
    analysis = {
        "Top 5 by Market Cap": top_5.to_dict(orient='records'),
        "Average Price of Top 50": avg_price,
        "Highest 24h Change": highest_change.to_dict(orient='records'),
        "Lowest 24h Change": lowest_change.to_dict(orient='records')
    }
    return df, analysis

def update_excel(df):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
    
    ws.delete_rows(1, ws.max_row)
    ws.append(["Name", "Symbol", "Current Price (USD)", "Market Cap", "24h Trading Volume", "24h Price Change (%)"])
    
    for row in dataframe_to_rows(df[['name', 'symbol', 'current_price', 'market_cap', 'total_volume', 'price_change_percentage_24h']], index=False, header=False):
        ws.append(row)
    
    wb.save(EXCEL_FILE)
    print("Excel updated successfully!")

def main():
    while True:
        data = fetch_crypto_data()
        if data:
            df, analysis = analyze_data(data)
            update_excel(df)
            print(analysis)
        time.sleep(UPDATE_INTERVAL)

if __name__ == "__main__":
    main()
